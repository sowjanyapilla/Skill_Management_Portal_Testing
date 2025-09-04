from fastapi import APIRouter, Depends, HTTPException, Query, Request, Response, Body, status, FastAPI, UploadFile, File, Form
from fastapi.responses import JSONResponse, HTMLResponse
from sqlalchemy import and_, or_, func, select, case, literal, desc, cast, Float, literal_column, distinct
from sqlalchemy.orm import selectinload, Session
from sqlalchemy.ext.asyncio import AsyncSession

# import os
# import json
# from docx import Document
# import PyPDF2
# from groq import Groq

from datetime import datetime
from typing import Optional, List, Dict
from collections import defaultdict
from fastapi.responses import StreamingResponse
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re
from fastapi.encoders import jsonable_encoder

from database import get_db
from models.skill import MasterSkill, SubSkill, EmployeeSkill, SkillStatus, EmployeeSkillHistory
from models.user import Employee
from schemas.user import EmployeeAuthenticated
from schemas.skill import (
    SkillCreate, SkillResponse, SubSkillResponse, EmployeeSkillUpdateRequest, EmployeeSkillResponse, PaginatedMasterSkills, MasterSkillSchema, SubSkillSchema, MasterSkillCreate, SubSkillCreate, SkillUpdate, SubSkillUpdate, MasterSkillSubSkillCreate, SubskillRequirement, MatchingRequest, ExportRequest
)
from routes.auth import get_current_user



router = APIRouter(prefix="/skills", tags=["Skills"])


# -------------------- Create a Skill with SubSkills --------------------
def normalize_subskill_name(name: str) -> str:
    """Trim and replace multiple spaces with single space, lowercase for consistency."""
    return re.sub(r'\s+', ' ', name.strip()).lower()

@router.post("/", response_model=SkillResponse, status_code=201)
async def create_skill(
    skill_data: SkillCreate,
    db: AsyncSession = Depends(get_db),
    current_user: EmployeeAuthenticated = Depends(get_current_user)
):
    # 1️⃣ Find or create MasterSkill
    result = await db.execute(select(MasterSkill).where(MasterSkill.skill_name == skill_data.skill_name))
    master_skill = result.scalars().first()
    if not master_skill:
        master_skill = MasterSkill(skill_name=skill_data.skill_name)
        db.add(master_skill)
        await db.flush()  # assign skill_id

    inserted_skills = []
    duplicate_skills = []
    employee_skills = []

    # Pre-fetch existing SubSkills under this MasterSkill
    result = await db.execute(select(SubSkill).where(SubSkill.skill_id == master_skill.skill_id))
    existing_sub_skills_db = {normalize_subskill_name(s.subskill_name): s for s in result.scalars().all()}

    # 2️⃣ Iterate over incoming subskills
    for sub_skill_data in skill_data.sub_skills:
        normalized_name = normalize_subskill_name(sub_skill_data.subskill_name)

        # Find or create SubSkill
        if normalized_name in existing_sub_skills_db:
            sub_skill = existing_sub_skills_db[normalized_name]
        else:
            sub_skill = SubSkill(
                skill_id=master_skill.skill_id,
                subskill_name=sub_skill_data.subskill_name.strip()
            )
            db.add(sub_skill)
            await db.flush()  # assign subskill_id
            existing_sub_skills_db[normalized_name] = sub_skill

        # 3️⃣ Check duplicate at EmployeeSkill level
        result = await db.execute(
            select(EmployeeSkill).where(
                EmployeeSkill.employee_id == current_user.id,
                EmployeeSkill.subskill_id == sub_skill.subskill_id
            )
        )
        existing_emp_skill = result.scalars().first()
        if existing_emp_skill:
            duplicate_skills.append(sub_skill.subskill_name.strip())
            continue

        # Fetch employee's approver
        result = await db.execute(
        select(Employee).where(Employee.id == current_user.id)
        )

        employee = result.scalars().first()
        approver_id = employee.approver_id if employee else None


        # 4️⃣ Insert EmployeeSkill for non-duplicate
        new_emp_skill = EmployeeSkill(
            employee_id=current_user.id,
            subskill_id=sub_skill.subskill_id,
            experience=sub_skill_data.experience_years,
            proficiency=sub_skill_data.employee_proficiency,
            certification=sub_skill_data.certification,
            certification_creation_date=sub_skill_data.certification_creation_date,
            certification_expiration_date=sub_skill_data.certification_expiration_date,
            status=SkillStatus.PENDING,
            approver_id=approver_id
            
        )
        db.add(new_emp_skill)
        await db.flush()  # get emp_skill_id

        employee_skills.append((new_emp_skill, sub_skill))
        inserted_skills.append(sub_skill.subskill_name.strip())

    await db.commit()

    # Refresh entities for response
    for emp_skill, sub_skill in employee_skills:
        await db.refresh(emp_skill)
    await db.refresh(master_skill)

    # 5️⃣ Build response with datetime serialization
    sub_skills_list = [
        SubSkillResponse(
            id=emp_skill.emp_skill_id,
            emp_skill_id=emp_skill.emp_skill_id,
            employee_id=emp_skill.employee_id,
            subskill_id=sub_skill.subskill_id,
            subskill_name=sub_skill.subskill_name,
            employee_proficiency=emp_skill.proficiency,
            manager_proficiency=None,
            experience_years=emp_skill.experience,
            certification=emp_skill.certification,
            certification_file_url=None,
            certification_creation_date=emp_skill.certification_creation_date.isoformat() if emp_skill.certification_creation_date else None,
            certification_expiration_date=emp_skill.certification_expiration_date.isoformat() if emp_skill.certification_expiration_date else None,
            status=emp_skill.status.value if emp_skill.status else None,
            manager_comments=emp_skill.manager_comments,
            approver_id=emp_skill.approver_id,
            created_at=emp_skill.created_date.isoformat() if emp_skill.created_date else None,
            last_updated_at=emp_skill.last_updated_at.isoformat() if getattr(emp_skill, "last_updated_at", None) else None
        )
        for emp_skill, sub_skill in employee_skills
    ]

    response = SkillResponse(
        id=master_skill.skill_id,
        user_id=current_user.id,
        skill_name=master_skill.skill_name,
        status=SkillStatus.PENDING.value,
        manager_comments=None,
        sub_skills=sub_skills_list
    )

    return JSONResponse(
    status_code=201,
    content=jsonable_encoder({
        "inserted_skills": inserted_skills,
        "duplicate_skills": duplicate_skills,
        "skill": response
    })
)


@router.get("/check-duplicates", response_model=List[dict])
async def check_duplicate_subskills(
    subskill_ids: List[int] = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: EmployeeAuthenticated = Depends(get_current_user)
):
    """
    Checks for existing EmployeeSkills based on a list of subskill IDs.
    Returns a list of subskill IDs that are already associated with the user.
    """
    # 1. Fetch existing EmployeeSkills for the current user
    result = await db.execute(
        select(EmployeeSkill.subskill_id)
        .where(
            EmployeeSkill.employee_id == current_user.id,
            EmployeeSkill.subskill_id.in_(subskill_ids)
        )
    )
    existing_subskill_ids = result.scalars().all()

    # 2. Return a list of dictionaries with the subskill name and existence status
    duplicate_info = []
    if existing_subskill_ids:
        subskill_result = await db.execute(
            select(SubSkill.subskill_id, SubSkill.subskill_name)
            .where(SubSkill.subskill_id.in_(existing_subskill_ids))
        )
        existing_subskills_map = {s.subskill_id: s.subskill_name for s in subskill_result.fetchall()}

        for sub_id in subskill_ids:
            if sub_id in existing_subskill_ids:
                duplicate_info.append({
                    "subskill_id": sub_id,
                    "subskill_name": existing_subskills_map.get(sub_id, "Unknown"),
                    "is_duplicate": True
                })
            else:
                duplicate_info.append({
                    "subskill_id": sub_id,
                    "is_duplicate": False
                })

    return duplicate_info

# -------------------- Get All Master Skills --------------------

@router.get("/my-skills")
async def get_my_skills(
    db: AsyncSession = Depends(get_db),
    current_user: EmployeeAuthenticated = Depends(get_current_user),
    skip: int = Query(0, ge=0),
    limit: int = Query(10, ge=1),
    status: Optional[str] = Query(None, description="PENDING | APPROVED | REJECTED"),
    skill_name: Optional[str] = Query(None, description="Filter by master skill name"),
    sub_skill_name: Optional[str] = Query(None, description="Filter by sub-skill name"),
):
    normalized_status = None
    if status:
        normalized_status = status.upper()
        if normalized_status not in {"PENDING", "APPROVED", "REJECTED"}:
            raise HTTPException(status_code=400, detail="Invalid status")

    # --------------------
    # 1. Choose base model and apply filters
    # --------------------
    if normalized_status == "REJECTED":
        base_query = (
            select(EmployeeSkillHistory)
            .options(selectinload(EmployeeSkillHistory.subskill).selectinload(SubSkill.master_skill))
            .where(
                EmployeeSkillHistory.employee_id == current_user.id,
                EmployeeSkillHistory.approval_status == "REJECTED"
            )
        )
        if skill_name:
            base_query = base_query.join(EmployeeSkillHistory.subskill).join(SubSkill.master_skill).where(MasterSkill.skill_name == skill_name)
        if sub_skill_name:
            base_query = base_query.join(EmployeeSkillHistory.subskill).where(SubSkill.subskill_name == sub_skill_name)
            
        result = await db.execute(base_query)
        all_rows = result.scalars().all()

        # Flatten all subskills
        flat_subskills = [
            {
                "master_skill_id": h.subskill.master_skill.skill_id,
                "master_skill_name": h.subskill.master_skill.skill_name,
                "emp_skill_id": h.emp_skill_id,
                "employee_id": h.employee_id,
                "subskill_id": h.subskill_id,
                "sub_skill_name": h.subskill.subskill_name,
                "proficiency_level": h.proficiency,
                "manager_proficiency": h.manager_proficiency,
                "status": h.approval_status.value if h.approval_status else None,
                "manager_comments": h.manager_comments,
                "experience_years": h.experience,
                "has_certification": bool(h.certification),
                "certification_file_url": h.certification,
                "created_at": h.created_at,
            }
            for h in all_rows
        ]

    else:
        base_query = (
            select(EmployeeSkill)
            .options(selectinload(EmployeeSkill.subskill).selectinload(SubSkill.master_skill))
            .where(EmployeeSkill.employee_id == current_user.id)
        )
        if normalized_status:
            base_query = base_query.where(EmployeeSkill.status == normalized_status)
        if skill_name:
            base_query = base_query.join(EmployeeSkill.subskill).join(SubSkill.master_skill).where(MasterSkill.skill_name == skill_name)
        if sub_skill_name:
            base_query = base_query.join(EmployeeSkill.subskill).where(SubSkill.subskill_name == sub_skill_name)

        result = await db.execute(base_query)
        all_rows = result.scalars().all()

        flat_subskills = [
            {
                "master_skill_id": es.subskill.master_skill.skill_id,
                "master_skill_name": es.subskill.master_skill.skill_name,
                "emp_skill_id": es.emp_skill_id,
                "employee_id": es.employee_id,
                "subskill_id": es.subskill_id,
                "sub_skill_name": es.subskill.subskill_name,
                "proficiency_level": es.proficiency,
                "manager_proficiency": None,
                "status": es.status.value if es.status else None,
                "manager_comments": es.manager_comments,
                "experience_years": es.experience,
                "has_certification": bool(es.certification),
                "certification_file_url": es.certification,
                "created_at": es.created_date,
            }
            for es in all_rows
        ]

    # --------------------
    # 2. Sort & Paginate at sub-skill level
    # --------------------
    flat_subskills.sort(key=lambda s: (s.get("master_skill_name", ""), s.get("sub_skill_name", "")))
    total_count = len(flat_subskills)
    paginated = flat_subskills[skip: skip + limit]

    # --------------------
    # 3. Group back into master skills
    # --------------------
    skills_data = []
    grouped = defaultdict(list)
    for sub in paginated:
        grouped[sub.get("master_skill_id")].append(sub)

    for ms_id, subs in grouped.items():
        skills_data.append({
            "id": ms_id,
            "user_id": current_user.id,
            "skill_name": subs[0].get("master_skill_name"),
            "sub_skills": subs,
            "status": subs[0].get("status"),
            "manager_comments": subs[0].get("manager_comments"),
            "created_at": subs[0].get("created_at"),
        })

    return {
        "skills": skills_data,
        "total": total_count,
        "skip": skip,
        "limit": limit,
    }


# new endpoint to fetch a single sub-skill
@router.get("/sub-skills/get/{subskill_id}")
async def get_single_sub_skill(
    subskill_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: EmployeeAuthenticated = Depends(get_current_user)
):
    try:
        # Check if the sub-skill exists for the current user in the main table
        result = await db.execute(
            select(EmployeeSkill)
            .options(selectinload(EmployeeSkill.subskill).selectinload(SubSkill.master_skill))
            .where(
                EmployeeSkill.subskill_id == subskill_id,
                EmployeeSkill.employee_id == current_user.id
            )
        )
        emp_skill = result.scalars().first()
        
        if emp_skill:
            return {
                "subskill_id": emp_skill.subskill_id,
                "sub_skill_name": emp_skill.subskill.subskill_name,
                "skill_name": emp_skill.subskill.master_skill.skill_name,
                "proficiency_level": emp_skill.proficiency,
                "experience_years": emp_skill.experience,
                "has_certification": bool(emp_skill.certification),
                "certification_file_url": emp_skill.certification,
                "certification_creation_date": emp_skill.certification_creation_date,
                "certification_expiration_date": emp_skill.certification_expiration_date,
                "status": emp_skill.status.value,
                "employee_id": emp_skill.employee_id,
                "emp_skill_id": emp_skill.emp_skill_id
            }

        # If not found in main table, check the history table for a rejected skill
        history_result = await db.execute(
            select(EmployeeSkillHistory)
            .options(selectinload(EmployeeSkillHistory.subskill).selectinload(SubSkill.master_skill))
            .where(
                EmployeeSkillHistory.subskill_id == subskill_id,
                EmployeeSkillHistory.employee_id == current_user.id,
                EmployeeSkillHistory.approval_status == "REJECTED"
            )
        )
        history_skill = history_result.scalars().first()

        if history_skill:
            return {
                "history_id": history_skill.history_id,
                "subskill_id": history_skill.subskill_id,
                "sub_skill_name": history_skill.subskill.subskill_name,
                "skill_name": history_skill.subskill.master_skill.skill_name,
                "proficiency_level": history_skill.proficiency,
                "experience_years": history_skill.experience,
                "has_certification": bool(history_skill.certification),
                "certification_file_url": history_skill.certification,
                "certification_creation_date": history_skill.certification_creation_date,
                "certification_expiration_date": history_skill.certification_expiration_date,
                "status": history_skill.approval_status.value,
                "employee_id": history_skill.employee_id,
                "emp_skill_id": history_skill.emp_skill_id
            }

        raise HTTPException(status_code=404, detail="Sub-skill not found for this user.")

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"An error occurred: {str(e)}")
    
# Get all master skills
@router.get("/master-skills")
async def get_master_skills(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(MasterSkill))
    skills = result.scalars().all()
    return [{"id": s.skill_id, "skill_name": s.skill_name} for s in skills]

@router.put("/master/{skill_id}")
async def update_master_skill(
    skill_id: int,
    skill_data: SkillUpdate,
    db: AsyncSession = Depends(get_db)
):
    # Fetch the skill
    result = await db.execute(select(MasterSkill).where(MasterSkill.skill_id == skill_id))
    skill = result.scalars().first()
    if not skill:
        raise HTTPException(status_code=404, detail="Skill not found")

    # Update skill name if provided
    if skill_data.skill_name is not None:
        skill.skill_name = skill_data.skill_name

    # Update sub-skills if provided
    if skill_data.sub_skills is not None:
        for sub in skill_data.sub_skills:
            result = await db.execute(
                select(SubSkill).where(
                    and_(
                        SubSkill.skill_id == skill.id,
                        SubSkill.sub_skill_name == sub.sub_skill_name
                    )
                )
            )
            sub_skill = result.scalars().first()
            if sub_skill:
                if sub.proficiency_level is not None:
                    sub_skill.proficiency_level = sub.proficiency_level

    await db.commit()
    return {"message": f"Skill {skill_data.skill_name} updated successfully"}

@router.put("/sub/bulk/{skill_id}")
async def update_master_skill_and_subskills(
    skill_id: int,
    payload: dict,
    db: AsyncSession = Depends(get_db)
):
    # 1. Fetch master skill
    result = await db.execute(
        select(MasterSkill).where(MasterSkill.skill_id == skill_id)
    )
    master_skill = result.scalars().first()
    if not master_skill:
        raise HTTPException(status_code=404, detail="Master skill not found")

    # 2. Update master skill name
    if "skill_name" in payload:
        master_skill.skill_name = payload["skill_name"]

    # 3. Add new subskills if provided
    if "new_subskills" in payload and isinstance(payload["new_subskills"], list):
        for sub in payload["new_subskills"]:
            new_sub = SubSkill(
                skill_id=skill_id,
                subskill_name=sub.get("subskill_name")
            )
            db.add(new_sub)

    await db.commit()
    return {"message": "Master skill and subskills updated successfully"}


@router.put("/sub/{subskill_id}")
async def update_sub_skill(
    subskill_id: int,
    subskill_data: SubSkillUpdate,
    db: AsyncSession = Depends(get_db)
):
    # Fetch the subskill
    result = await db.execute(
        select(SubSkill).where(SubSkill.subskill_id == subskill_id)
    )
    subskill = result.scalars().first()

    if not subskill:
        raise HTTPException(status_code=404, detail="Sub-skill not found")

    # Update name
    subskill.subskill_name = subskill_data.subskill_name

    await db.commit()
    await db.refresh(subskill)

    return {
        "message": f"Sub-skill {subskill.subskill_name} updated successfully",
        "subskill": {
            "subskill_id": subskill.subskill_id,
            "subskill_name": subskill.subskill_name,
            "skill_id": subskill.skill_id
        }
    }

# Get sub-skills by master skill
@router.get("/sub-skills/{master_skill_id}")
async def get_sub_skills(master_skill_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(SubSkill).where(SubSkill.skill_id == master_skill_id))
    sub_skills = result.scalars().all()
    return [{"id": s.subskill_id, "subskill_name": s.subskill_name} for s in sub_skills]

# --- Create MasterSkill with SubSkills ---
@router.post("/master/")
async def create_master_skill(
    payload: MasterSkillCreate, db: AsyncSession = Depends(get_db)
):
    # ✅ Check duplicate
    result = await db.execute(
        select(MasterSkill).where(MasterSkill.skill_name == payload.skill_name)
    )
    existing = result.scalars().first()
    if existing:
        raise HTTPException(status_code=400, detail="Master skill already exists.")

    master_skill = MasterSkill(skill_name=payload.skill_name)

    # ✅ Add sub-skills
    for sub in payload.sub_skills:
        sub_skill = SubSkill(subskill_name=sub.subskill_name)
        master_skill.subskills.append(sub_skill)

    db.add(master_skill)
    await db.commit()
    await db.refresh(master_skill)

    # ✅ Re-fetch with subskills eagerly loaded
    result = await db.execute(
        select(MasterSkill)
        .options(selectinload(MasterSkill.subskills))
        .where(MasterSkill.skill_id == master_skill.skill_id)
    )
    master_skill = result.scalars().first()

    return {
        "skill_id": master_skill.skill_id,
        "skill_name": master_skill.skill_name,
        "sub_skills": [
            {"subskill_id": s.subskill_id, "subskill_name": s.subskill_name}
            for s in master_skill.subskills
        ],
    }

@router.delete("/master/{skill_id}", status_code=200)
async def delete_skill(skill_id: int, db: AsyncSession = Depends(get_db)):
    # Fetch the skill
    result = await db.execute(select(MasterSkill).where(MasterSkill.skill_id == skill_id))
    skill = result.scalars().first()

    if not skill:
        raise HTTPException(status_code=404, detail="Skill not found")

    # Deleting skill will also delete sub-skills (if cascade is set in models)
    await db.delete(skill)
    await db.commit()

    return {"message": f"Skill {skill_id} and its sub-skills deleted successfully"}

@router.delete("/sub/{subskill_id}")
async def delete_subskill(
    subskill_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user)  # remove if no auth
):
    result = await db.execute(select(SubSkill).where(SubSkill.subskill_id == subskill_id))
    subskill = result.scalars().first()
    if not subskill:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"SubSkill with id {subskill_id} not found",
        )
    
    await db.delete(subskill)
    await db.commit()
    return {"message": f"Subskill {subskill.subskill_name} deleted successfully"}

    
def str_to_bool(val: Optional[str]):
    if val is None:
        return None
    if val.lower() in ["true", "1", "yes"]:
        return True
    elif val.lower() in ["false", "0", "no"]:
        return False
    return None

async def compute_skill_level_coverage(
    all_employee_skills: list,
    filter_skills: list
) -> float:
    """
    Compute skill coverage percentage for an employee.
    all_employee_skills: list of EmployeeSkill + SubSkill + MasterSkill tuples
    filter_skills: list of dicts with skill_name, min_proficiency, min_experience, max_experience, require_certification
    """

    total_skills = len(filter_skills)
    if total_skills == 0:
        return 0.0

    matched_skills = 0

    for f in filter_skills:
        skill_matches = [
            s for s in all_employee_skills
            if (
                f["skill_name"].lower() in s.MasterSkill.skill_name.lower()  # match master skill
                or f["skill_name"].lower() in s.SubSkill.subskill_name.lower()  # OR match sub-skill
            )
            and (f.get("min_proficiency") is None or (s.EmployeeSkill.proficiency is not None and s.EmployeeSkill.proficiency >= f["min_proficiency"]))
            and (f.get("min_experience") is None or (s.EmployeeSkill.experience is not None and s.EmployeeSkill.experience >= f["min_experience"]))
            and (f.get("max_experience") is None or (s.EmployeeSkill.experience is not None and s.EmployeeSkill.experience <= f["max_experience"]))
            and (not f.get("require_certification", False) or s.EmployeeSkill.certification is not None)
        ]
        if skill_matches:
            matched_skills += 1

    coverage = (matched_skills / total_skills) * 100
    return round(coverage, 2)

# This is the corrected version of your /matching endpoint.
# Copy and paste this to replace the existing one in your file.



# @router.get("/matching")
# async def get_employee_skill_coverage(
#     request: Request,
#     skill: Optional[str] = None,
#     proficiency: Optional[int] = None,
#     min_experience: Optional[float] = None,
#     max_experience: Optional[float] = None,
#     has_certification: Optional[bool] = None,
#     page: int = Query(1, ge=1),
#     page_size: int = Query(5, ge=1, le=100),
#     db: AsyncSession = Depends(get_db),
# ):
#     """
#     Returns employees grouped by skills with coverage %.
#     Pagination applied at employee level.
#     """

#     # Build filter list for coverage calculation
#     filter_skills = []
#     if skill:
#         skill_list = [s.strip() for s in skill.split(",") if s.strip()]
#         for s in skill_list:
#             filter_skills.append({
#                 "skill_name": s,
#                 "min_proficiency": proficiency,
#                 "min_experience": min_experience,
#                 "max_experience": max_experience,
#                 "require_certification": has_certification,
#             })

#     # Step 1: Build the filtered query for matching employees
#     filtered_emp_query = select(Employee).where(Employee.is_active == True)
    
#     # Check if ANY filter exists before adding joins and where clauses.
#     has_any_filter = skill or proficiency is not None or min_experience is not None or max_experience is not None or has_certification is not None

#     if has_any_filter:
#         # Explicitly join Employee to EmployeeSkill to resolve the ambiguity
#         filtered_emp_query = filtered_emp_query.join(EmployeeSkill, Employee.id == EmployeeSkill.employee_id)
        
#         # Now add the remaining joins
#         filtered_emp_query = filtered_emp_query.join(SubSkill).join(MasterSkill)
        
#         # This is the key change to handle comma-separated skills
#         if skill:
#             skill_list = [s.strip() for s in skill.split(",") if s.strip()]
            
#             # Build a list of individual OR conditions
#             skill_or_conditions = [
#                 or_(
#                     MasterSkill.skill_name.ilike(f"%{s}%"),
#                     SubSkill.subskill_name.ilike(f"%{s}%")
#                 ) for s in skill_list
#             ]
            
#             # Combine all OR conditions into one `where` clause
#             filtered_emp_query = filtered_emp_query.where(or_(*skill_or_conditions))

#         # Apply other filters to the joined tables
#         if proficiency is not None:
#             filtered_emp_query = filtered_emp_query.where(EmployeeSkill.proficiency >= proficiency)
#         if min_experience is not None:
#             filtered_emp_query = filtered_emp_query.where(EmployeeSkill.experience >= min_experience)
#         if max_experience is not None:
#             filtered_emp_query = filtered_emp_query.where(EmployeeSkill.experience <= max_experience)
#         if has_certification is not None:
#             if has_certification:
#                 filtered_emp_query = filtered_emp_query.where(EmployeeSkill.certification.isnot(None))
#             else:
#                 filtered_emp_query = filtered_emp_query.where(EmployeeSkill.certification.is_(None))

#     # Add a DISTINCT clause to ensure we only get unique employees,
#     # and GROUP BY to allow counting correctly.
#     filtered_emp_query = filtered_emp_query.group_by(Employee.id)

#     # Step 2: Get the total count of matching employees.
#     # We use a subquery to correctly count unique employees after joins and filters.
#     total_count_query = select(func.count()).select_from(filtered_emp_query.subquery())
#     total_matching_employees = (await db.execute(total_count_query)).scalar() or 0

#     # Step 3: Apply pagination to the filtered query and execute.
#     stmt = filtered_emp_query.offset((page - 1) * page_size).limit(page_size)
#     result = await db.execute(stmt)
#     employees = result.scalars().all()

#     # Step 4: Process the skills for this small, correctly paginated list.
#     employee_results = []
    
#     # The `filter_skills` list is already built correctly at the beginning.
    
#     for emp in employees:
#         # Fetch all skills + subskills + master skills for this specific employee.
#         # This part remains the same as it correctly fetches all related data for one employee.
#         stmt_skills = (
#             select(EmployeeSkill, SubSkill, MasterSkill)
#             .join(SubSkill, EmployeeSkill.subskill_id == SubSkill.subskill_id)
#             .join(MasterSkill, SubSkill.skill_id == MasterSkill.skill_id)
#             .filter(EmployeeSkill.employee_id == emp.id)
#         )
#         all_employee_skills = (await db.execute(stmt_skills)).all()

#         # Compute coverage
#         coverage = 0.0
#         if filter_skills:
#             coverage = await compute_skill_level_coverage(all_employee_skills, filter_skills)

#         skills_dict: dict[str, list] = {}

#         for emp_skill, subskill, master_skill in all_employee_skills:
#             key = master_skill.skill_name

#             # This inner loop also needs to handle the individual skills from the filter list.
#             if filter_skills:
#                 matched = False
#                 for f in filter_skills:
#                     if (
#                         f["skill_name"].lower() in master_skill.skill_name.lower()
#                         or f["skill_name"].lower() in subskill.subskill_name.lower()
#                     ) and (
#                         f.get("min_proficiency") is None or (emp_skill.proficiency or 0) >= f["min_proficiency"]
#                     ) and (
#                         f.get("min_experience") is None or (emp_skill.experience or 0) >= f["min_experience"]
#                     ) and (
#                         f.get("max_experience") is None or (emp_skill.experience or 0) <= f["max_experience"]
#                     ) and (
#                         f.get("require_certification") is None or (bool(emp_skill.certification) == f["require_certification"])
#                     ):
#                         matched = True
#                         break

#                 if not matched:
#                     continue

#             if key not in skills_dict:
#                 skills_dict[key] = []

#             skills_dict[key].append({
#                 "name": subskill.subskill_name,
#                 "proficiency": emp_skill.proficiency,
#                 "experience": emp_skill.experience,
#                 "hasCertification": bool(emp_skill.certification),
#                 "status": emp_skill.status.value if emp_skill.status else "PENDING",
#                 "certificationFile": emp_skill.certification,
#                 "certificationCreationDate": getattr(emp_skill, "certification_creation_date", None),
#                 "certificationExpirationDate": getattr(emp_skill, "certification_expiration_date", None),
#             })
        
#         if not skills_dict:
#              continue

#         skills_info = []
#         for skill_name, subskills in skills_dict.items():
#             matched = len([s for s in subskills if s["proficiency"] is not None and s["proficiency"] > 0])
#             skills_info.append({
#                 "skill_name": skill_name,
#                 "matched_subskills": matched,
#                 "total_subskills": len(subskills),
#                 "sub_skills": subskills,
#             })

#         employee_results.append({
#             "employee_id": emp.emp_id,
#             "employee_name": emp.name,
#             "coverage": coverage,
#             "skills": skills_info,
#         })
    
#     total_pages = (total_matching_employees + page_size - 1) // page_size

#     return {
#         "results": employee_results,
#         "page": page,
#         "page_size": page_size,
#         "total_pages": total_pages,
#         "total_employees": total_matching_employees,
#     }


@router.post("/matching")
async def get_employee_skill_coverage(
    payload: MatchingRequest = Body(...),
    page: int = Query(1, ge=1),
    page_size: int = Query(5, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
):
    """
    Accepts JSON body: { "requirements":[{subskill_id, min_experience, max_experience, min_proficiency, require_certification}, ...] }
    Returns paginated list of employees with score, coverage and matched skills (aggregated in DB).
    """

    # --- Normalize incoming requirements ---
    reqs = payload.requirements or []
    N = len(reqs)

    # --- Build per-requirement SQL conditions (only SubSkill matches) ---
    matched_row_conditions = []
    for r in reqs:
        conds = [EmployeeSkill.subskill_id == r.subskill_id]
        if r.min_proficiency is not None:
            conds.append(EmployeeSkill.proficiency >= r.min_proficiency)
        if r.min_experience is not None:
            conds.append(EmployeeSkill.experience >= r.min_experience)
        if r.max_experience is not None:
            conds.append(EmployeeSkill.experience <= r.max_experience)
        if r.require_certification is not None:
            if r.require_certification:
                conds.append(EmployeeSkill.certification.isnot(None))
            else:
                conds.append(EmployeeSkill.certification.is_(None))

        matched_row_conditions.append(and_(*conds))

    # --- Score expression (per matched row) ---
    exp_norm = cast(func.coalesce(EmployeeSkill.experience, 0), Float) / literal(132.0)
    prof_norm = cast(func.coalesce(EmployeeSkill.proficiency, 0), Float) / literal(5.0)
    cert_val = case((EmployeeSkill.certification.isnot(None), literal(1.0)), else_=literal(0.0))

    subskill_score_expr = (literal(0.8) * exp_norm) + (literal(0.15) * prof_norm) + (literal(0.05) * cert_val)
    sum_subskill_scores = func.sum(subskill_score_expr)

    if N > 0:
        final_score_expr = (func.coalesce(sum_subskill_scores, literal(0.0)) / literal(float(N))) * literal(100.0)
    else:
        final_score_expr = literal(0.0)

    # Build aggregated JSON of matched rows
    skills_json_expr = func.coalesce(
        func.jsonb_agg(
            func.jsonb_build_object(
                "subskill_id", SubSkill.subskill_id,
                "master_skill", MasterSkill.skill_name,
                "subskill", SubSkill.subskill_name,
                "proficiency", EmployeeSkill.proficiency,
                "experience", EmployeeSkill.experience,
                "hasCertification", case((EmployeeSkill.certification.isnot(None), literal(True)), else_=literal(False)),
                "status", EmployeeSkill.status
            )
        ),
        literal_column("'[]'::jsonb")
    ).label("skills")

    matched_count_expr = func.count(distinct(SubSkill.subskill_id)).label("matched_count")

    # Base aggregated query
    filtered_emp_query = (
        select(
            Employee,
            final_score_expr.label("final_score"),
            matched_count_expr,
            skills_json_expr
        )
        .select_from(Employee)
        .join(EmployeeSkill, Employee.id == EmployeeSkill.employee_id)
        .join(SubSkill, EmployeeSkill.subskill_id == SubSkill.subskill_id)
        .join(MasterSkill, SubSkill.skill_id == MasterSkill.skill_id)
        .where(Employee.is_active == True)
    )

    if matched_row_conditions:
        filtered_emp_query = filtered_emp_query.where(or_(*matched_row_conditions))

    filtered_emp_query = filtered_emp_query.group_by(Employee.id)

    # Total count for pagination
    total_count_query = select(func.count()).select_from(filtered_emp_query.subquery())
    total_matching_employees = (await db.execute(total_count_query)).scalar() or 0

    # Paginate & order
    stmt = filtered_emp_query.order_by(desc(literal_column("final_score"))).offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(stmt)
    rows = result.all()  # (Employee, final_score, matched_count, skills_json)

    employee_results = []
    for emp, final_score, matched_count, skills_json in rows:
        coverage = 0.0
        if N > 0:
            coverage = float(matched_count or 0) / float(N) * 100.0

        skills_list = skills_json if skills_json is not None else []

        # Build grouped skills structure (by master skill)
        skills_dict: dict[str, list] = {}
        for s in skills_list:
            master = s.get("master_skill") or "Unknown"
            skills_dict.setdefault(master, []).append({
                "subskill_id": s.get("subskill_id"),
                "name": s.get("subskill"),
                "proficiency": s.get("proficiency"),
                "experience": s.get("experience"),
                "hasCertification": bool(s.get("hasCertification")),
                "status": s.get("status"),
            })

        skills_info = []
        for skill_name, subskills in skills_dict.items():
            matched = len([x for x in subskills if x["proficiency"] is not None and x["proficiency"] > 0])
            skills_info.append({
                "skill_name": skill_name,
                "matched_subskills": matched,
                "total_subskills": len(subskills),
                "sub_skills": subskills,
            })

        # Note: adapt these attribute names to your model (emp.emp_id / emp.name) if needed
        employee_results.append({
            "employee_id": getattr(emp, "emp_id", getattr(emp, "id", None)),
            "employee_name": getattr(emp, "name", getattr(emp, "full_name", None)),
            "coverage": round(coverage, 2),
            "score": round(float(final_score or 0.0), 2),
            "skills": skills_info,
        })

    total_pages = (total_matching_employees + page_size - 1) // page_size

    return {
        "results": employee_results,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "total_employees": total_matching_employees,
    }


# Simple GET /subskills endpoint used by the frontend
# @router.get("/subskills")
# async def get_all_subskills(session: AsyncSession = Depends(get_db)):
#     result = await session.execute(
#         select(SubSkill.subskill_id, SubSkill.subskill_name)
#     )
#     rows = result.all()
#     # rows will be list of Row tuples; map to objects
#     return [{"id": r.subskill_id, "name": r.subskill_name} for r in rows]

@router.get("/subskills")
async def get_all_subskills(session: AsyncSession = Depends(get_db)):
    """
    Return list of subskills with their master skill name.
    Response: [{ "id": int, "name": str, "master_skill": str | None }, ...]
    """
    stmt = (
        select(
            SubSkill.subskill_id.label("id"),
            SubSkill.subskill_name.label("name"),
            MasterSkill.skill_name.label("master_skill"),
        )
        .select_from(SubSkill)
        .join(MasterSkill, SubSkill.skill_id == MasterSkill.skill_id, isouter=True)
    )

    result = await session.execute(stmt)
    rows = result.all()

    return [
        {
            "id": r.id,
            "name": r.name,
            "master_skill": r.master_skill,  # may be None if no master found
        }
        for r in rows
    ]



# # -------------------- export skills matching users -------------------- 
# @router.get("/matching/export")
# async def export_employee_skills(
#     skill: str | None = None,
#     proficiency: int | None = None,
#     min_experience: float | None = None,
#     max_experience: float | None = None,
#     has_certification: bool | None = None,
#     db: AsyncSession = Depends(get_db),
# ):
#     """
#     Export all employees (not paginated) with their grouped skills into Excel.
#     Applies main filters + table header filters.
#     """

#     # Build filter list
#     filter_skills = []
#     skill_list = []
#     if skill:
#         skill_list = [s.strip() for s in skill.split(",") if s.strip()]
#         for s in skill_list:
#             filter_skills.append({
#                 "skill_name": s,
#                 "min_proficiency": proficiency,
#                 "min_experience": min_experience,
#                 "max_experience": max_experience,
#                 "require_certification": has_certification,
#             })

#     # ✅ Fetch all employees (no pagination here)
#     stmt = select(Employee).where(Employee.is_active == True)
#     result = await db.execute(stmt)
#     employees = result.scalars().all()

#     # Excel workbook
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Employee Skills"

#     # Header row
#     headers = [
#         "Employee Name", "Employee ID",
#         "Skill", "Sub-skill",
#         "Proficiency", "Experience (yrs)",
#         "Certification", "Status",
#         "Coverage (%)"
#     ]
#     ws.append(headers)

#     # Style headers
#     for col in range(1, len(headers)+1):
#         cell = ws.cell(row=1, column=col)
#         cell.font = Font(bold=True)
#         cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
#         cell.alignment = Alignment(horizontal="center")

#     # Fill data
#     for emp in employees:
#         stmt_skills = (
#             select(EmployeeSkill, SubSkill, MasterSkill)
#             .join(SubSkill, EmployeeSkill.subskill_id == SubSkill.subskill_id)
#             .join(MasterSkill, SubSkill.skill_id == MasterSkill.skill_id)
#             .filter(EmployeeSkill.employee_id == emp.id)
#         )
#         all_employee_skills = (await db.execute(stmt_skills)).all()

#         if not all_employee_skills:
#             continue

#         # Compute coverage
#         coverage = 0.0
#         if filter_skills:
#             coverage = await compute_skill_level_coverage(all_employee_skills, filter_skills)

#         skills_dict = {}
#         for emp_skill, subskill, master_skill in all_employee_skills:
#             key = master_skill.skill_name

#             if key not in skills_dict:
#                 skills_dict[key] = []

#             skills_dict[key].append({
#                 "name": subskill.subskill_name,
#                 "proficiency": emp_skill.proficiency,
#                 "experience": emp_skill.experience,
#                 "hasCertification": bool(emp_skill.certification),
#                 "status": emp_skill.status.value if emp_skill.status else "PENDING",
#             })

#         # Skip employees with no matching subskills
#         if not skills_dict:
#             continue

#         first_row_for_employee = True
#         for skill_name, subskills in skills_dict.items():
#             first_row_for_skill = True
#             for sub in subskills:
#                 row = [
#                     emp.name if first_row_for_employee else "",
#                     emp.emp_id if first_row_for_employee else "",
#                     skill_name if first_row_for_skill else "",
#                     sub["name"],
#                     sub["proficiency"],
#                     sub["experience"],
#                     "Certified" if sub["hasCertification"] else "Not Certified",
#                     sub["status"],
#                     coverage if first_row_for_employee else ""
#                 ]
#                 ws.append(row)
#                 first_row_for_employee = False
#                 first_row_for_skill = False

#         # Add empty row between employees
#         ws.append([])

#     # Save to BytesIO
#     file_stream = BytesIO()
#     wb.save(file_stream)
#     file_stream.seek(0)

#     return StreamingResponse(
#         file_stream,
#         media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#         headers={"Content-Disposition": "attachment; filename=employee_skills.xlsx"}
#     )


@router.post("/matching/export")
async def export_employee_skills(
    payload: ExportRequest = Body(...),
    db: AsyncSession = Depends(get_db),
):
    """
    POST body: { "requirements": [ { subskill_id, min_experience, max_experience, min_proficiency, require_certification }, ... ] }
    Exports Excel for all matching employees (no pagination).
    """

    reqs = payload.requirements or []
    N = len(reqs)

    # --- Build per-requirement matched-row conditions (only SubSkill matches) ---
    matched_row_conditions = []
    for r in reqs:
        conds = [EmployeeSkill.subskill_id == r.subskill_id]
        if r.min_proficiency is not None:
            conds.append(EmployeeSkill.proficiency >= r.min_proficiency)
        if r.min_experience is not None:
            conds.append(EmployeeSkill.experience >= r.min_experience)
        if r.max_experience is not None:
            conds.append(EmployeeSkill.experience <= r.max_experience)
        if r.require_certification is not None:
            if r.require_certification:
                conds.append(EmployeeSkill.certification.isnot(None))
            else:
                conds.append(EmployeeSkill.certification.is_(None))
        matched_row_conditions.append(and_(*conds))

    # --- Score expression (per matched row) ---
    exp_norm = cast(func.coalesce(EmployeeSkill.experience, 0), Float) / literal(132.0)
    prof_norm = cast(func.coalesce(EmployeeSkill.proficiency, 0), Float) / literal(5.0)
    cert_val = case((EmployeeSkill.certification.isnot(None), literal(1.0)), else_=literal(0.0))

    subskill_score_expr = (literal(0.8) * exp_norm) + (literal(0.15) * prof_norm) + (literal(0.05) * cert_val)
    sum_subskill_scores = func.sum(subskill_score_expr)

    if N > 0:
        final_score_expr = (func.coalesce(sum_subskill_scores, literal(0.0)) / literal(float(N))) * literal(100.0)
    else:
        final_score_expr = literal(0.0)

    # aggregate matched rows into JSON (Postgres jsonb)
    skills_json_expr = func.coalesce(
        func.jsonb_agg(
            func.jsonb_build_object(
                "subskill_id", SubSkill.subskill_id,
                "master_skill", MasterSkill.skill_name,
                "subskill", SubSkill.subskill_name,
                "proficiency", EmployeeSkill.proficiency,
                "experience", EmployeeSkill.experience,
                "hasCertification", case((EmployeeSkill.certification.isnot(None), literal(True)), else_=literal(False)),
                "status", EmployeeSkill.status
            )
        ),
        literal_column("'[]'::jsonb")
    ).label("skills")

    matched_count_expr = func.count(distinct(SubSkill.subskill_id)).label("matched_count")

    # --- Build aggregated query: if no requirements -> aggregate ALL skills per employee
    filtered_emp_query = (
        select(
            Employee,
            final_score_expr.label("final_score"),
            matched_count_expr,
            skills_json_expr
        )
        .select_from(Employee)
        .join(EmployeeSkill, Employee.id == EmployeeSkill.employee_id)
        .join(SubSkill, EmployeeSkill.subskill_id == SubSkill.subskill_id)
        .join(MasterSkill, SubSkill.skill_id == MasterSkill.skill_id)
        .where(Employee.is_active == True)
    )

    # If we have matched_row_conditions, restrict aggregated rows to only those matched rows (so skills_json contains only matched rows).
    # If there are no requirements, we keep all rows (so skills_json contains all skills).
    if matched_row_conditions:
        filtered_emp_query = filtered_emp_query.where(or_(*matched_row_conditions))

    filtered_emp_query = filtered_emp_query.group_by(Employee.id)

    # Order by score desc for export readability
    stmt = filtered_emp_query.order_by(desc(literal_column("final_score")))
    result = await db.execute(stmt)
    rows = result.all()  # each row = (Employee, final_score, matched_count, skills_json)

    # --- Excel workbook setup ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Skills"

    headers = [
        "Employee Name", "Employee ID", "Skill", "Sub-skill",
        "Proficiency", "Experience (yrs)", "Certification", "Status",
        "Coverage (%)", "Score"
    ]
    ws.append(headers)

    # header style
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Fill rows
    for row in rows:
        emp, final_score, matched_count, skills_json = row
        # coverage calculation
        coverage = 0.0
        if N > 0:
            coverage = float(matched_count or 0) / float(N) * 100.0
        else:
            coverage = 0.0

        skills_list = skills_json if skills_json is not None else []

        # group skills by master skill for nicer Excel layout
        skills_dict: dict[str, list] = {}
        for s in skills_list:
            master = s.get("master_skill") or "Unknown"
            skills_dict.setdefault(master, []).append({
                "name": s.get("subskill"),
                "proficiency": s.get("proficiency"),
                "experience": s.get("experience"),
                "hasCertification": bool(s.get("hasCertification")),
                "status": s.get("status"),
            })

        if not skills_dict:
            # still append a row with employee and empty skill, if you want to include employees with no matched skills
            ws.append([getattr(emp, "name", ""), getattr(emp, "emp_id", ""), "", "", "", "", "", "", round(coverage,2), round(float(final_score or 0.0),2)])
            continue

        first_row_for_employee = True
        for skill_name, subskills in skills_dict.items():
            first_row_for_skill = True
            for sub in subskills:
                excel_row = [
                    getattr(emp, "name", "") if first_row_for_employee else "",
                    getattr(emp, "emp_id", "") if first_row_for_employee else "",
                    skill_name if first_row_for_skill else "",
                    sub["name"],
                    sub["proficiency"],
                    sub["experience"],
                    "Certified" if sub["hasCertification"] else "Not Certified",
                    sub["status"],
                    round(coverage, 2) if first_row_for_employee else "",
                    round(float(final_score or 0.0), 2) if first_row_for_employee else ""
                ]
                ws.append(excel_row)
                first_row_for_employee = False
                first_row_for_skill = False

        # spacer row between employees
        ws.append([])

    # Auto-width columns (simple heuristic)
    for i, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
            except Exception:
                val = ""
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(max(10, max_len + 2), 60)

    # Save workbook to BytesIO and return as streaming response
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=employee_skills.xlsx"}
    )

@router.get("/skills/get", response_model=SubSkillResponse)
async def get_subskill(employee_id: int, subskill_id: int, db: AsyncSession = Depends(get_db)):
    # Fetch EmployeeSkill with subskill eagerly loaded
    result = await db.execute(
        select(EmployeeSkill)
        .options(selectinload(EmployeeSkill.subskill))
        .where(
            EmployeeSkill.employee_id == employee_id,
            EmployeeSkill.subskill_id == subskill_id
        )
    )
    skill = result.scalars().first()

    if not skill:
        raise HTTPException(status_code=404, detail="Skill not found")

    return SubSkillResponse(
        id=skill.emp_skill_id,
        emp_skill_id=skill.emp_skill_id,
        employee_id=skill.employee_id,
        subskill_id=skill.subskill_id,
        subskill_name=skill.subskill.subskill_name if skill.subskill else "Unknown",
        employee_proficiency=skill.proficiency,
        manager_proficiency=None,  # or fetch from history if needed
        experience_years=skill.experience,
        certification=skill.certification,
        certification_file_url=None,  # if you have file URLs
        certification_creation_date=skill.certification_creation_date,
        certification_expiration_date=skill.certification_expiration_date,
        status=skill.status.value if skill.status else None,
        manager_comments=skill.manager_comments,
        approver_id=skill.approver_id,
        created_at=skill.created_date or datetime.utcnow(),
        last_updated_at=getattr(skill, "updated_at", datetime.utcnow())
    )

@router.put("/update")
async def update_employee_skill(request: EmployeeSkillUpdateRequest, db: AsyncSession = Depends(get_db)):

    # Fetch the existing record by employee_id + subskill_id
    result = await db.execute(
        select(EmployeeSkill).where(
            EmployeeSkill.employee_id == request.employee_id,
            EmployeeSkill.subskill_id == request.subskill_id
        )
    )
    skill_record = result.scalars().first()

    if not skill_record:
        raise HTTPException(status_code=404, detail="Employee skill not found")

    # Update fields
    skill_record.experience = request.experience_years
    skill_record.proficiency = request.proficiency_level
   
    skill_record.certification_file_url = getattr(request, "certification_file_url", None)
    skill_record.certification_creation_date = request.certification_creation_date
    skill_record.certification_expiration_date = request.certification_expiration_date
    skill_record.status = SkillStatus.PENDING  # Reset status
    skill_record.manager_comments = None

    # Commit and refresh
    await db.commit()
    await db.refresh(skill_record)

    return {
        "message": "Employee skill updated successfully",
        "data": {
            "emp_skill_id": skill_record.emp_skill_id,
            "employee_id": skill_record.employee_id,
            "subskill_id": skill_record.subskill_id,
            "employee_proficiency": skill_record.proficiency,
            "experience_years": skill_record.experience,
            "has_certification": skill_record.certification,
            "certification_file_url": skill_record.certification_file_url,
            "certification_creation_date": skill_record.certification_creation_date,
            "certification_expiration_date": skill_record.certification_expiration_date,
            "status": skill_record.status.value
        }
    }


# @router.get("/history/{history_id}")
# def get_skill_history(history_id: int, db: Session = Depends(get_db)):
#     history_record = db.query(EmployeeSkillHistory).filter(EmployeeSkillHistory.history_id == history_id).first()
#     if not history_record:
#         raise HTTPException(status_code=404, detail="History record not found")
    
#     # Optional: serialize data for frontend
#     return {
#         "history_id": history_record.history_id,
#         "employee_id": history_record.employee_id,
#         "subskill_id": history_record.subskill_id,
#         "experience": history_record.experience,
#         "proficiency": history_record.proficiency,
#         "certification": history_record.certification,
#         "certification_creation_date": history_record.certification_creation_date,
#         "certification_expiration_date": history_record.certification_expiration_date,
#         "manager_proficiency": history_record.manager_proficiency,
#         "manager_comments": history_record.manager_comments,
#         "approval_status": history_record.approval_status.value
#     }

@router.get("/master/all")
async def get_all_master_skills(db: AsyncSession = Depends(get_db)):
    """
    Fetch all master skills with their subskills (AsyncSession + eager loading)
    """
    # Use selectinload to fetch related subskills
    stmt = select(MasterSkill).options(selectinload(MasterSkill.subskills))
    result = await db.execute(stmt)
    master_skills = result.scalars().unique().all()

    response = []
    for skill in master_skills:
        subskills_list = [
            {"subskill_id": sub.subskill_id, "subskill_name": sub.subskill_name, "skill_id": skill.skill_id}
            for sub in skill.subskills
        ]
        response.append({
            "skill_id": skill.skill_id,
            "skill_name": skill.skill_name,
            "subskills": subskills_list
        })

    return {"master_skills": response}


@router.get("/all")
async def get_all_skills_pg(session: AsyncSession = Depends(get_db)):
    """
    Returns dict of skill_name -> list of subskill names.
    Example:
    {
      "Salesforce": ["Salesforce Admin", "Salesforce Development", "LWC"],
      "Power BI": ["Pivot", "DAX", "Power Query"]
    }
    """
    sub_json = func.coalesce(
        func.jsonb_agg(SubSkill.subskill_name)
            .filter(SubSkill.subskill_id.isnot(None)),
        literal_column("'[]'::jsonb")
    ).label("subskills")

    stmt = (
        select(
            MasterSkill.skill_name.label("master_name"),
            sub_json
        )
        .select_from(MasterSkill)
        .outerjoin(SubSkill, SubSkill.skill_id == MasterSkill.skill_id)
        .group_by(MasterSkill.skill_id, MasterSkill.skill_name)
        .order_by(MasterSkill.skill_name)
    )

    result = await session.execute(stmt)
    rows = result.all()

    # Convert rows -> dict {skill_name: [subskills]}
    out = {r.master_name: r.subskills for r in rows}
    return out


import json
import re
from fastapi import FastAPI, UploadFile, File, HTTPException, Form
from fastapi.responses import HTMLResponse, JSONResponse
from docx import Document
import PyPDF2
from groq import Groq


# Hardcoded API key - replace with your actual key
GROQ_API_KEY = "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"
client = Groq(api_key=GROQ_API_KEY)


def extract_text_from_pdf(file):
    reader = PyPDF2.PdfReader(file)
    text = ""
    for page in reader.pages:
        text += page.extract_text() or ""
    return text

def extract_text_from_docx(uploaded_file):
    # uploaded_file is a SpooledTemporaryFile (from UploadFile.file)
    content = uploaded_file.read()   # read binary content
    uploaded_file.seek(0)            # reset pointer for safety
    doc = Document(BytesIO(content)) # wrap in BytesIO
    return "\n".join([p.text for p in doc.paragraphs])


def extract_text(file, filename):
    if filename.lower().endswith(".pdf"):
        return extract_text_from_pdf(file)
    elif filename.lower().endswith(".docx"):
        return extract_text_from_docx(file)
    return None

def parse_jd_with_groq(text, skills_data: dict):
    prompt = f"""
You are an information extraction engine.
Input: Job Description text.
Output: STRICT JSON only, no explanation.

You are provided with an allowed skills dictionary.
You must only pick skills and subskills from that dictionary.

⚠️ CRITICAL RULES:
1. "skill" and "sub_skill" MUST always be filled with valid values from the dictionary.
   - Never return null for either field.
   - If no match exists, place the entire item under "unmatched_skills".

2. If the JD mentions only a broad skill (e.g., "Salesforce 10+ years") without subskills:
   - Expand this into ALL subskills of that skill from the dictionary.
   - Apply the same experience and proficiency to all subskills.

3. If the JD mentions an ambiguous subskill that exists under multiple skills
   (e.g., "Pivot" exists in Excel and Power BI):
   - Use the surrounding context in the JD to decide the correct parent skill.
   - If the context is unclear, default to the most common/likely parent (based on industry usage).
   - Never assign the same subskill under multiple skills unless the JD explicitly mentions both.

4. If the JD provides only general experience (e.g., "5+ years experience in Salesforce"):
   - Apply that min/max experience to all subskills of Salesforce.
   - Do not leave subskills with null experience.

5. If the JD mentions proficiency levels (beginner, intermediate, expert):
   - Map them to numbers (1–5), where 1 = beginner, 3 = intermediate, 5 = expert.

6. If the JD specifies certifications:
   - Mark "required_certification" = true.
   - Otherwise, always null.

7. The final JSON MUST follow this structure:

{{
  "job_title":"string — concise role title extracted or synthesized from the JD (e.g., \"Senior Salesforce Developer\"). If the JD has no explicit title, create a short descriptive title using the main skill(s) and inferred seniority (e.g., \"Salesforce Developer — Senior\").",
  "structured_data": [
    {{
      "skill": "string (from provided dictionary, never null)",
      "sub_skill": "string (from provided dictionary, never null)",
      "min_proficiency": int or null (1–5),
      "min_experience": float or null (years),
      "max_experience": float or null (years),
      "required_certification": true or null
    }}
  ],
  "unmatched_skills": ["list of skills/subskills not found in dictionary e.g.,"Java(exp 2-3, prof 3, cert yes)","C++(exp 5+, prof 4, cert no)" "]
}}
📘 Dictionary to use for validation:
{json.dumps(skills_data, indent=2)}

📘 Experience handling rules:
"2-4 years" → min=2, max=4.
"2 years experience" → min=2, max=2.
"X+ years" → min=X, max=null.
No mention → both null.

📘 Output rules:
Return ONLY valid JSON, no explanations.
Every structured_data entry MUST have a skill and sub_skill from the dictionary.

Job Description:
{text}
"""
    try:
        response = client.chat.completions.create(
            model="gemma2-9b-it",
            messages=[{"role": "user", "content": prompt}],
            temperature=0,
            max_tokens=2048,
        )
        raw_output = response.choices[0].message.content.strip()
        match = re.search(r"\{[\s\S]*\}", raw_output)
        print(raw_output)
        if match:
            return json.loads(match.group(0))
        else:
            return {"error": "No JSON found", "raw_output": raw_output}
    except Exception as e:
        return {"error": str(e)}



@router.post("/upload_jd")
async def upload_jd(
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_db),
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="Empty filename")

    # 1. Fetch skills dict directly from DB
    skills_dict = await get_all_skills_pg(session)

    # 2. Extract text
    text = extract_text(file.file, file.filename)
    if not text:
        raise HTTPException(status_code=400, detail="Unsupported file format")

    # 3. Call LLM parser with skills_dict
    result = parse_jd_with_groq(text, skills_dict)

    return JSONResponse(content=result)